#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
学生情報管理
"""

import os
import csv
from utils.file_utils import detect_excel_format, check_drm_protection
from utils.crypto import ENCRYPTION_AVAILABLE, encrypt_data, decrypt_data


ENCRYPTED_STUDENTS_FILE = "students_encrypted.dat"


def load_students_info():
    """学生情報を読み込み"""
    
    # まず暗号化キャッシュを確認
    if ENCRYPTION_AVAILABLE and os.path.exists(ENCRYPTED_STUDENTS_FILE):
        print(f"🔓 暗号化された学生情報を読み込み中...")
        try:
            with open(ENCRYPTED_STUDENTS_FILE, 'rb') as f:
                encrypted_data = f.read()
            cached_data = decrypt_data(encrypted_data)
            if cached_data:
                print(f"✅ 暗号化キャッシュから{len(cached_data)}人の学生情報を読み込みました")
                return cached_data
            else:
                print(f"⚠️  暗号化キャッシュの読み込みに失敗しました。元ファイルから再読み込みします")
        except:
            pass
    
    # ファイルをチェック
    possible_files = ["students.csv", "students.xlsx"]
    
    found_file = None
    for filename in possible_files:
        if os.path.exists(filename):
            found_file = filename
            break
    
    if not found_file:
        print(f"INFO: Student information file not found")
        print(f"   (one of: {', '.join(possible_files)})")
        return None
    
    # DRM保護をチェック
    if check_drm_protection(found_file):
        print(f"="*60)
        print(f"🔒 {found_file} はDRM保護されています")
        print(f"="*60)
        print(f"⚠️  このファイルはDRM(デジタル著作権管理)で保護されており、")
        print(f"    プログラムで自動的に読み込むことができません。\n")
        print(f"📋 対処方法:")
        print(f"   1. Excelで {found_file} を開く")
        print(f"   2. 「ファイル」→「名前を付けて保存」")
        print(f"   3. 形式を「CSV UTF-8 (*.csv)」に変更")
        print(f"   4. ファイル名を「students.csv」として保存")
        print(f"   5. このスクリプトと同じフォルダに配置")
        print(f"   6. スクリプトを再実行")
        print(f"="*60)
        print()
        return None
    
    # 実際の形式を検出
    actual_format = detect_excel_format(found_file)
    
    print(f"Reading {found_file}... (format: {actual_format.upper()})")
    
    try:
        students = {}
        
        # CSV形式の場合
        if actual_format == 'csv':
            with open(found_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                header = next(reader)
                
                num_columns = len(header)
                name_column_index = 9 if num_columns >= 10 else 2
                
                print(f"   検出: {num_columns}列 → 学生名は{'J列' if name_column_index == 9 else 'C列'}から読み込み")
                
                row_count = 0
                for row_idx, row in enumerate(reader, start=2):
                    if len(row) > name_column_index and row[name_column_index]:
                        attendance_number = row[0]
                        class_code = row[1]
                        student_name = row[name_column_index]
                        
                        if attendance_number and class_code and student_name:
                            name_key = str(student_name).replace(' ', '').replace('　', '').strip()
                            
                            student_info = {
                                'attendance_number': str(attendance_number).strip(),
                                'class_code': str(class_code).strip(),
                                'student_name': str(student_name).strip(),
                                'name_key': name_key
                            }
                            
                            # 既に同じ名前の学生がいる場合
                            if name_key in students:
                                existing = students[name_key]
                                if not isinstance(existing, list):
                                    students[name_key] = [existing]
                                    print(f"📋 複数クラス検出: {student_name}")
                                
                                if not any(s['class_code'] == class_code for s in students[name_key]):
                                    students[name_key].append(student_info)
                                    print(f"   └ {class_code} を追加 (計{len(students[name_key])}個)")
                            else:
                                students[name_key] = student_info
                            
                            row_count += 1
        
        # 新しいExcel形式(.xlsx)
        elif actual_format == 'xlsx':
            try:
                from openpyxl import load_workbook
            except ImportError:
                print("="*60)
                print("⚠️  openpyxlがインストールされていません")
                print()
                print("📋 推奨: CSVファイルを使用してください")
                print("   1. Excelでファイルを開く")
                print("   2. 「ファイル」→「名前を付けて保存」")
                print("   3. 形式を「CSV UTF-8 (*.csv)」に変更")
                print("   4. ファイル名を「students.csv」として保存")
                print()
                print("または、openpyxlをインストール:")
                print("   pip install openpyxl --break-system-packages")
                print("="*60)
                return None
            
            wb = load_workbook(found_file, read_only=True, data_only=True)
            ws = wb.active
            
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            num_columns = len([c for c in first_row if c is not None])
            name_column_index = 9 if num_columns >= 10 else 2
            
            print(f"   検出: {num_columns}列 → 学生名は{'J列' if name_column_index == 9 else 'C列'}から読み込み")
            
            row_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) > name_column_index and row[name_column_index]:
                    attendance_number = row[0]
                    class_code = row[1]
                    student_name = row[name_column_index]
                    
                    if attendance_number and class_code and student_name:
                        name_key = str(student_name).replace(' ', '').replace('　', '').strip()
                        
                        student_info = {
                            'attendance_number': str(attendance_number).strip(),
                            'class_code': str(class_code).strip(),
                            'student_name': str(student_name).strip(),
                            'name_key': name_key
                        }
                        
                        if name_key in students:
                            existing = students[name_key]
                            if not isinstance(existing, list):
                                students[name_key] = [existing]
                                print(f"📋 複数クラス検出: {student_name}")
                            
                            if not any(s['class_code'] == class_code for s in students[name_key]):
                                students[name_key].append(student_info)
                                print(f"   └ {class_code} を追加 (計{len(students[name_key])}個)")
                        else:
                            students[name_key] = student_info
                        
                        row_count += 1
            
            wb.close()
        
        # 古いExcel形式(.xls) - サポート終了
        else:
            print("="*60)
            print("⚠️  古いExcel形式(.xls)は非対応です")
            print()
            print("📋 対処方法:")
            print(f"   1. Excelで {found_file} を開く")
            print(f"   2. 「ファイル」→「名前を付けて保存」")
            print(f"   3. 形式を「CSV UTF-8 (*.csv)」に変更")
            print(f"   4. ファイル名を「students.csv」として保存")
            print(f"   5. このスクリプトと同じフォルダに配置")
            print(f"   6. スクリプトを再実行")
            print("="*60)
            return None
        
        if students:
            print(f"✅ {row_count}人の学生情報を読み込みました")
            
            # 複数クラス記号を持つ学生をログ出力
            multi_class_students = {k: v for k, v in students.items() if isinstance(v, list)}
            if multi_class_students:
                print(f"📋 複数クラス記号を持つ学生: {len(multi_class_students)}人")
                for name_key, info_list in multi_class_students.items():
                    codes = [item['class_code'] for item in info_list]
                    print(f"   - {info_list[0]['student_name']}: {', '.join(codes)}")
            
            # 暗号化キャッシュに保存
            if ENCRYPTION_AVAILABLE:
                encrypted_data = encrypt_data(students)
                if encrypted_data:
                    try:
                        with open(ENCRYPTED_STUDENTS_FILE, 'wb') as f:
                            f.write(encrypted_data)
                        print(f"🔐 学生情報を暗号化キャッシュに保存しました")
                    except:
                        print(f"⚠️  暗号化キャッシュの保存に失敗しました(通常動作に影響はありません)")
            
            return students
        else:
            print(f"⚠️  学生情報が見つかりませんでした(データが空)")
            return None
        
    except Exception as e:
        print(f"❌ ファイル読み込みエラー:")
        print(f"   {type(e).__name__}: {e}")
        print()
        print(f"💡 対処方法:")
        print(f"   1. Excelで {found_file} を開く")
        print(f"   2. 「ファイル」→「名前を付けて保存」")
        print(f"   3. 形式を「CSV UTF-8 (*.csv)」に変更")
        print(f"   4. ファイル名を「students.csv」として保存")
        print(f"   5. このスクリプトと同じフォルダに配置")
        return None
